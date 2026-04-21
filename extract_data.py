#!/usr/bin/env python3
"""Extract data from Excel files into JSON for the FED dashboard."""
import json
import os
import openpyxl
from datetime import datetime

DOWNLOADS = os.path.expanduser("~/Downloads")
OUTPUT = os.path.join(os.path.dirname(__file__), "src", "data")
os.makedirs(OUTPUT, exist_ok=True)

# Normalize inconsistent teacher name spellings from Excel
TEACHER_NAME_FIXES = {
    "lasse hægland": "Lasse Hægland",
    "lasse høgland": "Lasse Hægland",
    "lasse hægland": "Lasse Hægland",
    "adrian d souza": "Adrian D Souza",
}

def normalize_teacher(name):
    if not name:
        return None
    stripped = name.strip()
    lookup = stripped.lower()
    if lookup in TEACHER_NAME_FIXES:
        return TEACHER_NAME_FIXES[lookup]
    return stripped

ABBREV_TO_NAME = {
    "INTRO": "Introduction",
    "DES": "Design",
    "HTML": "HTML and CSS",
    "PME": "Project Methodology",
    "SP1": "Semester Project 1",
    "JS1": "JavaScript 1",
    "AGC1": "Agency 1",
    "PE1": "Exam Project 1",
    "POR1": "Portfolio 1",
    "JS2": "JavaScript 2",
    "WFL": "Workflow",
    "CSS": "CSS Frameworks",
    "SP2": "Semester Project 2",
    "DVP": "Development Platforms",
    "JSF": "JavaScript Frameworks",
    "AGC2": "Agency 2",
    "PE2": "Exam Project 2",
    "POR2": "Portfolio 2",
    # Legacy / older study plans
    "CMS": "Content Management Systems",
    "IND": "Industry Knowledge",
    "INDES": "Interaction Design",
    "PORT1": "Portfolio 1",
}

ABBREV_TO_CODE = {
    "INTRO": None,
    "DES": "FM1AIDE10",
    "HTML": "FM1AIHC10",
    "PME": "FM1AIPM25",
    "SP1": "FM1AIS175",
    "JS1": "FM1AIJ110",
    "AGC1": "FM1AIA110",
    "PE1": "FM1AIP175",
    "POR1": "FM1AIPO25",
    "JS2": "FM2AJJ210",
    "WFL": "FM2AJWF05",
    "CSS": "FM2AJCF05",
    "SP2": "FM2AJS210",
    "DVP": "FM2AJDP05",
    "JSF": "FM2AJJF75",
    "AGC2": "FM2AJA205",
    "PE2": "FM2AJP210",
    "POR2": "FM2AJPO25",
}

# ─── 1. COURSES + TEACHERS ───────────────────────────────────────────────

def extract_courses_and_teachers():
    wb = openpyxl.load_workbook(
        os.path.join(DOWNLOADS, "Teacher course responsibility-2025-revised.xlsx"),
        data_only=True,
    )
    ws = wb["Sheet1"]

    courses = []
    teachers_set = set()

    # Year 1 courses: rows 20-28
    # Year 2 courses: rows 32-41
    row_ranges = [(20, 28, 1), (32, 41, 2)]

    for start_row, end_row, year in row_ranges:
        for r in range(start_row, end_row + 1):
            code = ws.cell(r, 2).value
            name = ws.cell(r, 3).value
            semester = ws.cell(r, 4).value
            weeks = ws.cell(r, 5).value
            hours = ws.cell(r, 6).value
            credits = ws.cell(r, 7).value
            responsible = ws.cell(r, 8).value
            second = ws.cell(r, 9).value
            third = ws.cell(r, 10).value

            if not name:
                continue

            # Course names in Excel are like "Design(DES)" — extract abbreviation from parentheses
            import re
            name_str = name.strip()
            paren_match = re.search(r'\(([^)]+)\)', name_str)
            if paren_match:
                abbrev = paren_match.group(1).strip().replace(" ", "")
                clean_name = re.sub(r'\s*\([^)]+\)', '', name_str).strip()
            else:
                abbrev = None
                clean_name = name_str
                for ab, full_name in ABBREV_TO_NAME.items():
                    if full_name.lower() == clean_name.lower():
                        abbrev = ab
                        break

            course = {
                "code": code if code else None,
                "name": clean_name,
                "abbreviation": abbrev or clean_name[:4].upper(),
                "year": year,
                "semester": int(semester) if semester else None,
                "weeksFT": int(weeks) if weeks else 1,
                "hours": int(hours) if hours else None,
                "credits": float(credits) if credits else None,
                "responsibleTeacher": normalize_teacher(responsible),
                "secondTeacher": normalize_teacher(second),
                "thirdTeacher": normalize_teacher(third),
            }
            courses.append(course)

            for t in [responsible, second, third]:
                normalized = normalize_teacher(t)
                if normalized:
                    teachers_set.add(normalized)

    # Teacher summary: rows 44-51
    teachers = []
    for r in range(44, 52):
        tname = ws.cell(r, 2).value
        if not tname or not tname.strip():
            continue
        normalized = normalize_teacher(tname)
        if not normalized or normalized.lower() == "teacher name":
            continue
        resp_count = ws.cell(r, 3).value or 0
        second_count = ws.cell(r, 4).value or 0
        third_count = ws.cell(r, 5).value or 0
        on_leave = ws.cell(r, 6).value
        back = ws.cell(r, 7).value

        teachers.append({
            "name": normalized,
            "coursesResponsible": int(resp_count) if isinstance(resp_count, (int, float)) else 0,
            "courses2nd": int(second_count) if isinstance(second_count, (int, float)) else 0,
            "courses3rd": int(third_count) if isinstance(third_count, (int, float)) else 0,
            "onLeave": str(on_leave) if on_leave else None,
            "backDate": str(back) if back else None,
        })

    # Add Shabana Jahan if not in summary table (she's only on Introduction)
    teacher_names = {t["name"] for t in teachers}
    for t in teachers_set:
        if t not in teacher_names and t.lower() not in {n.lower() for n in teacher_names}:
            teachers.append({
                "name": t,
                "coursesResponsible": 0,
                "courses2nd": 0,
                "courses3rd": 0,
                "onLeave": None,
                "backDate": None,
            })

    return courses, teachers


# ─── 2. INTAKES + SCHEDULE ───────────────────────────────────────────────

def extract_intakes_and_schedule():
    wb_data = openpyxl.load_workbook(
        os.path.join(DOWNLOADS, "FED Access List.xlsx"), data_only=True
    )
    wb_fmt = openpyxl.load_workbook(
        os.path.join(DOWNLOADS, "FED Access List.xlsx")
    )
    ws_data = wb_data["AUG2022 FED"]
    ws_fmt = wb_fmt["AUG2022 FED"]

    intakes = []
    schedule = []

    # Row 14 has intake IDs, row 10 has student counts, row 11 has study plans, row 12 has LMS
    for c in range(4, 45):
        intake_id = ws_data.cell(14, c).value
        if not intake_id or not intake_id.strip():
            continue

        student_count = ws_data.cell(10, c).value
        if not student_count or int(student_count) == 0:
            continue

        study_plan = ws_data.cell(11, c).value
        lms = ws_data.cell(12, c).value

        # Parse intake type (FT/PT) and start month/year
        iid = intake_id.strip()
        is_ft = "FT" in iid
        is_pt = "PT" in iid

        # Extract month and year: e.g. AUG24FT -> AUG, 2024
        import re
        match = re.match(r"([A-Z]+)(\d{2})(FT|PT)", iid)
        if match:
            month_str = match.group(1)
            year_2d = int(match.group(2))
            start_year = 2000 + year_2d
            intake_type = match.group(3)
        else:
            month_str = iid[:3]
            start_year = 2020
            intake_type = "PT" if is_pt else "FT"

        intakes.append({
            "id": iid,
            "type": intake_type,
            "startMonth": month_str,
            "startYear": start_year,
            "studentCount": int(student_count),
            "studyPlan": study_plan.strip() if study_plan else None,
            "lms": lms.strip() if lms else None,
        })

        # Extract weekly schedule for this intake
        for r in range(17, ws_data.max_row + 1):
            week_date = ws_data.cell(r, 2).value
            cal_week = ws_data.cell(r, 3).value
            cell_val = ws_data.cell(r, c).value

            if not week_date:
                continue
            if not cell_val:
                continue

            val_str = str(cell_val).strip()

            # Parse "N.ABBREV" format
            parts = val_str.split(".", 1)
            if len(parts) == 2:
                week_num_str = parts[0].strip()
                abbrev = parts[1].strip()
                try:
                    course_week_num = int(week_num_str)
                except ValueError:
                    course_week_num = 0
            else:
                course_week_num = 0
                abbrev = val_str

            # Handle compound entries like "POR2 / 0.INTRO"
            if "/" in abbrev:
                abbrev = abbrev.split("/")[0].strip()

            # Detect colors from formatted workbook
            fmt_cell = ws_fmt.cell(r, c)
            fill = fmt_cell.fill
            fg = fill.fgColor

            is_assignment = False
            is_course_start = False

            if fg:
                try:
                    theme = fg.theme
                    if theme == 4:  # Blue theme = assignment week
                        is_assignment = True
                except:
                    pass
                try:
                    rgb = fg.rgb
                    if isinstance(rgb, str) and "FFC000" in rgb:  # Orange = course start
                        is_course_start = True
                except:
                    pass

            # Format date
            if isinstance(week_date, datetime):
                date_str = week_date.strftime("%Y-%m-%d")
            else:
                date_str = str(week_date)[:10]

            schedule.append({
                "intake": iid,
                "weekDate": date_str,
                "calendarWeek": str(cal_week) if cal_week else "",
                "courseAbbrev": abbrev,
                "courseWeekNumber": course_week_num,
                "isAssignmentWeek": is_assignment,
                "isCourseStart": is_course_start,
            })

    return intakes, schedule


# ─── MAIN ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Extracting courses and teachers...")
    courses, teachers = extract_courses_and_teachers()

    print("Extracting intakes and schedule...")
    intakes, schedule = extract_intakes_and_schedule()

    # Write JSON files
    with open(os.path.join(OUTPUT, "courses.json"), "w") as f:
        json.dump(courses, f, indent=2, ensure_ascii=False)
    print(f"  courses.json: {len(courses)} courses")

    with open(os.path.join(OUTPUT, "teachers.json"), "w") as f:
        json.dump(teachers, f, indent=2, ensure_ascii=False)
    print(f"  teachers.json: {len(teachers)} teachers")

    with open(os.path.join(OUTPUT, "intakes.json"), "w") as f:
        json.dump(intakes, f, indent=2, ensure_ascii=False)
    print(f"  intakes.json: {len(intakes)} intakes")

    with open(os.path.join(OUTPUT, "schedule.json"), "w") as f:
        json.dump(schedule, f, indent=2, ensure_ascii=False)
    print(f"  schedule.json: {len(schedule)} weekly entries")

    # Also write the abbreviation map
    with open(os.path.join(OUTPUT, "abbreviations.json"), "w") as f:
        json.dump(ABBREV_TO_NAME, f, indent=2, ensure_ascii=False)

    print("Done!")
